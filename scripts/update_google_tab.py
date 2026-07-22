"""Baixa a planilha do Google Sheets e atualiza a aba dados-google da base.

Usado pelo workflow .github/workflows/atualiza-dados-google.yml.
Substitui integralmente a aba `dados-google` de dados/base_atual.xlsx pelo
conteúdo da aba homônima da planilha compartilhada (fonte alimentada pelo
script do Google Ads no MCC). As demais abas são preservadas.

Só regrava o arquivo quando o conteúdo realmente mudou, para evitar commits
vazios. Variável de ambiente LOCAL_SRC permite testar com um arquivo local.
"""

from __future__ import annotations

import io
import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SHEET_ID = "1mAI4ICmOw6SDxIHLmwnT5SkzYYF57k7Y0vI0WmtKV0k"
TAB = "dados-google"
BASE_PATH = Path("dados/base_atual.xlsx")
EXPECTED_HEADER = [
    "Marca", "Campanha", "Tipo de campanha", "Cliques", "Impr.",
    "Custo", "CPC", "Início", "Fim", "Semana",
]
EXPORT_URL = (
    "https://docs.google.com/spreadsheets/d/" + SHEET_ID + "/export?format=xlsx"
)


def fail(message: str) -> None:
    print("ERRO: " + message)
    sys.exit(1)


def download_source() -> bytes:
    local = os.environ.get("LOCAL_SRC")
    if local:
        return Path(local).read_bytes()
    import requests

    response = requests.get(EXPORT_URL, timeout=180)
    if response.status_code != 200:
        fail(f"Download da planilha retornou HTTP {response.status_code}.")
    content_type = response.headers.get("Content-Type", "")
    if "html" in content_type:
        fail(
            "O Google devolveu uma página em vez do arquivo — verifique se a "
            "planilha está compartilhada como 'qualquer pessoa com o link'."
        )
    return response.content


def read_rows(payload: bytes) -> list[tuple]:
    workbook = load_workbook(io.BytesIO(payload), data_only=True)
    if TAB not in workbook.sheetnames:
        fail(f"A planilha baixada não tem a aba '{TAB}'.")
    sheet = workbook[TAB]
    rows = []
    for row in sheet.iter_rows(values_only=True):
        values = tuple((row + (None,) * 10)[:10])
        rows.append(values)
    while rows and all(v is None or v == "" for v in rows[-1]):
        rows.pop()
    if not rows:
        fail("A aba baixada está vazia.")
    header = [str(v).strip() if v is not None else "" for v in rows[0]]
    if header != EXPECTED_HEADER:
        fail(f"Cabeçalho inesperado na planilha: {header}")
    if len(rows) < 2:
        fail("A aba baixada não tem linhas de dados.")
    return rows


def normalized(value):
    """Compara datas e números de forma estável."""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, float) and value == int(value):
        return int(value)
    if value == "":
        return None
    return value


def current_rows(workbook) -> list[tuple]:
    sheet = workbook[TAB]
    rows = [tuple((row + (None,) * 10)[:10]) for row in sheet.iter_rows(values_only=True)]
    while rows and all(v is None or v == "" for v in rows[-1]):
        rows.pop()
    return rows


def main() -> None:
    payload = download_source()
    new_rows = read_rows(payload)

    if not BASE_PATH.exists():
        fail(f"{BASE_PATH} não encontrado no repositório.")
    base = load_workbook(BASE_PATH)
    if TAB not in base.sheetnames:
        fail(f"{BASE_PATH} não tem a aba '{TAB}'.")

    old = [[normalized(v) for v in row] for row in current_rows(base)]
    new = [[normalized(v) for v in row] for row in new_rows]
    if old == new:
        print(f"Sem mudanças na aba {TAB} ({len(new) - 1} linhas de dados).")
        return

    index = base.sheetnames.index(TAB)
    del base[TAB]
    sheet = base.create_sheet(TAB, index)
    for row in new_rows:
        sheet.append(list(row))
    for row_cells in sheet.iter_rows(min_row=2, min_col=8, max_col=9):
        for cell in row_cells:
            cell.number_format = "dd/mm/yyyy"
    base.save(BASE_PATH)
    print(
        f"Aba {TAB} atualizada: {len(old) - 1 if old else 0} → {len(new) - 1} "
        "linhas de dados."
    )


if __name__ == "__main__":
    main()
