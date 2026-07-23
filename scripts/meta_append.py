"""Acrescenta linhas diárias do Meta Ads na aba dados-face da base.

Entrada: um arquivo CSV (separador ;) com o cabeçalho:
data;pagina;campanha;tipo_resultado;resultados;alcance;impressoes;cliques;gasto

- data em YYYY-MM-DD; resultados pode ser vazio (quando o Meta não informa);
- números podem vir como "3.724", "1.234,56", "R$203,47 BRL" — são normalizados;
- Custo por resultado é calculado (gasto / resultados);
- Início = Término = data; Semana = período do mês (1 a 7 / 8 a 14 / 15 a 21 / 22 a fim);
- Deduplicação: linhas cuja (página, data) já existem na aba são ignoradas,
  então o mesmo arquivo pode ser processado duas vezes sem duplicar.

Uso: python scripts/meta_append.py caminho/do/arquivo.csv
"""

from __future__ import annotations

import csv
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

BASE_PATH = Path("dados/base_atual.xlsx")
TAB = "dados-face"
N_COLS = 13  # até "Nome do conjunto de anúncios"

EXPECTED_HEADER = [
    "data", "pagina", "campanha", "tipo_resultado", "resultados",
    "alcance", "impressoes", "cliques", "gasto",
]


def parse_number(raw: str):
    """Converte "3.724", "1.234,56", "R$203,47 BRL" em número (ou None)."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    text = re.sub(r"[^\d.,-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif text.count(".") == 1 and len(text.split(".")[1]) == 3:
        # "3.724" é separador de milhar
        text = text.replace(".", "")
    elif text.count(".") > 1:
        text = text.replace(".", "")
    value = float(text)
    return int(value) if value == int(value) else round(value, 2)


def semana_label(date: datetime) -> str:
    day = date.day
    if day <= 7:
        return "1 a 7"
    if day <= 14:
        return "8 a 14"
    if day <= 21:
        return "15 a 21"
    if date.month == 12:
        last = 31
    else:
        last = (date.replace(day=1, month=date.month % 12 + 1) - timedelta(days=1)).day
    return f"22 a {last}"


def existing_keys(sheet):
    keys = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        pagina, inicio = row[0], row[9]
        if pagina and isinstance(inicio, datetime):
            keys.add((str(pagina).strip(), inicio.strftime("%Y-%m-%d")))
    return keys


def last_data_row(sheet) -> int:
    for idx in range(sheet.max_row, 0, -1):
        if sheet.cell(row=idx, column=1).value not in (None, ""):
            return idx
    return 1


def main() -> None:
    if len(sys.argv) != 2:
        print("Uso: python scripts/meta_append.py arquivo.csv")
        sys.exit(1)
    csv_path = Path(sys.argv[1])
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=";")
        rows = list(reader)
    if not rows or [c.strip().lower() for c in rows[0]] != EXPECTED_HEADER:
        print(f"ERRO: cabeçalho esperado: {';'.join(EXPECTED_HEADER)}")
        sys.exit(1)

    base = load_workbook(BASE_PATH)
    sheet = base[TAB]
    keys = existing_keys(sheet)

    new_rows = []
    skipped = 0
    for line_number, row in enumerate(rows[1:], start=2):
        if not any(cell.strip() for cell in row):
            continue
        if len(row) < 9:
            print(f"ERRO: linha {line_number} tem {len(row)} colunas (esperado 9).")
            sys.exit(1)
        data_str, pagina, campanha, tipo, resultados, alcance, impr, cliques, gasto = (
            cell.strip() for cell in row[:9]
        )
        date = datetime.strptime(data_str, "%Y-%m-%d")
        if (pagina, data_str) in keys:
            skipped += 1
            continue
        resultados_n = parse_number(resultados)
        gasto_n = parse_number(gasto) or 0
        custo_por_resultado = (
            round(gasto_n / resultados_n, 6) if resultados_n else None
        )
        new_rows.append([
            pagina,
            campanha,
            tipo or None,
            resultados_n,
            parse_number(alcance),
            parse_number(impr),
            custo_por_resultado,
            parse_number(cliques),
            gasto_n,
            date,
            date,
            semana_label(date),
            None,
        ])

    if not new_rows:
        print(f"Nada novo a gravar ({skipped} linha(s) já existiam).")
        return

    start = last_data_row(sheet) + 1
    for offset, values in enumerate(new_rows):
        for col, value in enumerate(values, start=1):
            sheet.cell(row=start + offset, column=col, value=value)
    for row_cells in sheet.iter_rows(
        min_row=start, max_row=start + len(new_rows) - 1, min_col=10, max_col=11
    ):
        for cell in row_cells:
            cell.number_format = "dd/mm/yyyy"
    base.save(BASE_PATH)
    print(
        f"Gravadas {len(new_rows)} linha(s) novas na {TAB} "
        f"(ignoradas {skipped} duplicadas)."
    )


if __name__ == "__main__":
    main()
